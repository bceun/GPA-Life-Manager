from datetime import datetime
from html import escape
from io import BytesIO
from itertools import product

import joblib
import numpy as np
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# =========================================================
# 기본 설정
# =========================================================
st.set_page_config(
    page_title="Campus Twin",
    page_icon="🎓",
    layout="wide",
)

MODEL_TARGET_MIN = 0.0
MODEL_TARGET_MAX = 2.009058
DISPLAY_SCORE_MAX = 4.5
EPSILON = 1e-8

DAYS = ["월요일", "화요일", "수요일", "목요일", "금요일", "토요일", "일요일"]

FEATURE_COLS = [
    "study_hours",
    "attendance",
    "sleep_hours",
    "screen_time",
    "physical_activity",
    "stress",
]

FEATURE_LABELS = {
    "study_hours": "공부 시간",
    "attendance": "출석률",
    "sleep_hours": "수면 시간",
    "screen_time": "스크린타임",
    "physical_activity": "신체활동",
    "stress": "스트레스",
}

FEATURE_UNITS = {
    "study_hours": "시간",
    "attendance": "%",
    "sleep_hours": "시간",
    "screen_time": "시간",
    "physical_activity": "시간",
    "stress": "단계",
}

LABEL_TO_FEATURE = {label: feature for feature, label in FEATURE_LABELS.items()}

SENSITIVITY_SETTINGS = {
    "study_hours": (1.0, 0.0, 12.0, "공부 시간 1시간 증가"),
    "attendance": (5.0, 0.0, 100.0, "출석률 5% 증가"),
    "sleep_hours": (1.0, 0.0, 12.0, "수면 시간 1시간 증가"),
    "screen_time": (-1.0, 0.0, 16.0, "스크린타임 1시간 감소"),
    "physical_activity": (1.0, 0.0, 8.0, "신체활동 1시간 증가"),
    "stress": (-1.0, 1.0, 10.0, "스트레스 1단계 감소"),
}


# =========================================================
# 화면 스타일
# =========================================================
st.markdown(
    """
    <style>
    .recommendation-card {
        border: 1px solid #d9dee8;
        border-radius: 14px;
        padding: 20px;
        min-height: 420px;
        background: #ffffff;
        margin-bottom: 10px;
        overflow-wrap: break-word;
    }
    .recommendation-card-selected {
        border: 3px solid #ff4b4b;
        border-radius: 14px;
        padding: 18px;
        min-height: 420px;
        background: #fff5f5;
        box-shadow: 0 4px 14px rgba(255, 75, 75, 0.16);
        margin-bottom: 10px;
        overflow-wrap: break-word;
    }
    .card-title {font-size: 1.45rem; font-weight: 700; margin-bottom: 14px;}
    .card-score-label {font-size: .95rem; margin-bottom: 4px;}
    .card-score {font-size: 1.8rem; font-weight: 700; margin-bottom: 8px;}
    .card-delta-positive {color: #11823b; font-weight: 700; margin-bottom: 14px;}
    .card-reason {
        background: #eaf3ff;
        border-radius: 9px;
        padding: 12px;
        margin: 12px 0 10px;
    }
    .selected-badge {
        display: inline-block;
        padding: 4px 9px;
        margin-bottom: 10px;
        border-radius: 20px;
        background: #ff4b4b;
        color: white;
        font-size: .85rem;
        font-weight: 700;
    }
    .difficulty-low, .difficulty-medium, .difficulty-high {
        border-radius: 8px;
        padding: 14px;
        margin-top: 10px;
    }
    .difficulty-low {background: #eaf8ef; border-left: 5px solid #2ca25f;}
    .difficulty-medium {background: #fff8df; border-left: 5px solid #e6a700;}
    .difficulty-high {background: #fff0f0; border-left: 5px solid #d94343;}
    .step-card {
        border: 1px solid #d9dee8;
        border-radius: 12px;
        padding: 16px;
        background: #fafbfc;
        min-height: 230px;
    }
    .progress-summary {
        text-align: center;
        font-size: 1.15rem;
        font-weight: 700;
        padding: 10px;
    }
    </style>
    """,
    unsafe_allow_html=True,
)


# =========================================================
# 모델
# =========================================================
@st.cache_resource
def load_models():
    return joblib.load("lr_model.pkl"), joblib.load("scaler.pkl")


try:
    lr_model, scaler = load_models()
except Exception as error:
    st.error("모델 파일을 불러오지 못했습니다. lr_model.pkl과 scaler.pkl을 확인해 주세요.")
    st.exception(error)
    st.stop()


def make_input_df(values):
    missing = [column for column in FEATURE_COLS if column not in values]
    if missing:
        raise ValueError(f"필요한 입력 컬럼이 없습니다: {missing}")
    return pd.DataFrame([values])[FEATURE_COLS]


def convert_to_display_score(raw_score):
    clipped = np.clip(raw_score, MODEL_TARGET_MIN, MODEL_TARGET_MAX)
    converted = (
        (clipped - MODEL_TARGET_MIN)
        / (MODEL_TARGET_MAX - MODEL_TARGET_MIN)
        * DISPLAY_SCORE_MAX
    )
    return float(np.clip(converted, 0.0, DISPLAY_SCORE_MAX))


def predict_from_values(values):
    input_df = make_input_df(values)
    scaled = scaler.transform(input_df)
    prediction = lr_model.predict(scaled)
    raw_score = float(np.asarray(prediction).reshape(-1)[0])
    raw_score = float(np.clip(raw_score, MODEL_TARGET_MIN, MODEL_TARGET_MAX))
    return {
        "raw_score": raw_score,
        "display_score": convert_to_display_score(raw_score),
    }


def calculate_total_hours(values):
    return (
        values["study_hours"]
        + values["sleep_hours"]
        + values["screen_time"]
        + values["physical_activity"]
    )


def create_range(start, stop, step):
    if stop < start:
        return np.array([float(start)])
    return np.round(np.arange(start, stop + step / 2, step), 2)


# =========================================================
# 분석
# =========================================================
def analyze_sensitivity(values):
    base = predict_from_values(values)
    rows = []

    for feature, (change, min_value, max_value, description) in SENSITIVITY_SETTINGS.items():
        changed = values.copy()
        changed[feature] = float(np.clip(changed[feature] + change, min_value, max_value))
        prediction = predict_from_values(changed)

        rows.append(
            {
                "생활요인 변화": description,
                "현재 값": values[feature],
                "변경 값": changed[feature],
                "변경 후 환산 점수": prediction["display_score"],
                "환산 점수 변화량": prediction["display_score"] - base["display_score"],
                "원본 점수 변화량": prediction["raw_score"] - base["raw_score"],
            }
        )

    return pd.DataFrame(rows).sort_values("환산 점수 변화량", ascending=False).reset_index(drop=True)


def compare_plans(current, future):
    current_prediction = predict_from_values(current)
    future_prediction = predict_from_values(future)

    rows = [
        {
            "생활요인": FEATURE_LABELS[feature],
            "현재": current[feature],
            "미래 계획": future[feature],
            "변화": future[feature] - current[feature],
            "단위": FEATURE_UNITS[feature],
        }
        for feature in FEATURE_COLS
    ]

    return {
        "current_raw_score": current_prediction["raw_score"],
        "future_raw_score": future_prediction["raw_score"],
        "current_display_score": current_prediction["display_score"],
        "future_display_score": future_prediction["display_score"],
        "display_score_change": (
            future_prediction["display_score"] - current_prediction["display_score"]
        ),
        "comparison_df": pd.DataFrame(rows),
    }


# =========================================================
# 추천 알고리즘
# =========================================================
def calculate_change_cost(current, candidate):
    denominators = {
        "study_hours": 12.0,
        "attendance": 100.0,
        "sleep_hours": 12.0,
        "screen_time": 16.0,
        "physical_activity": 8.0,
        "stress": 9.0,
    }
    return float(
        sum(abs(candidate[f] - current[f]) / denominators[f] for f in FEATURE_COLS)
    )


def describe_change_cost(cost):
    if cost < 0.08:
        return "낮음"
    if cost < 0.18:
        return "보통"
    return "높음"


def calculate_difficulty_score(cost):
    return int(np.clip(np.ceil(cost * 12), 1, 5))


def describe_difficulty(score):
    return {1: "매우 쉬움", 2: "쉬움", 3: "보통", 4: "어려움", 5: "매우 어려움"}[score]


def get_difficulty_message(score):
    if score <= 2:
        return (
            "difficulty-low",
            "무리 없이 시작할 수 있는 계획입니다.",
            "현재 생활과의 차이가 크지 않아 이번 주부터 시작해도 부담이 비교적 적습니다.",
        )
    if score == 3:
        return (
            "difficulty-medium",
            "일정 조정이 필요한 계획입니다.",
            "한 번에 모든 목표를 적용하기보다 공부·수면·출석 계획을 단계적으로 조정하세요.",
        )
    return (
        "difficulty-high",
        "단계적으로 실행할 것을 권장합니다.",
        "생활 변화 폭이 큽니다. 3주 단계계획을 따라 목표 수준에 천천히 도달하세요.",
    )


def has_any_change(current, candidate):
    return any(abs(float(current[f]) - float(candidate[f])) >= EPSILON for f in FEATURE_COLS)


def generate_candidate_values(
    current,
    changeable_features,
    max_study,
    min_sleep,
    max_screen_reduction,
    max_attendance_increase,
):
    study = (
        create_range(current["study_hours"], min(max_study, current["study_hours"] + 2.0, 12.0), 0.5)
        if "study_hours" in changeable_features
        else np.array([current["study_hours"]])
    )

    if "sleep_hours" in changeable_features:
        sleep_start = max(current["sleep_hours"], min_sleep)
        sleep_stop = min(12.0, current["sleep_hours"] + 1.5)
        sleep = (
            create_range(sleep_start, sleep_stop, 0.5)
            if sleep_start <= sleep_stop
            else np.array([current["sleep_hours"]])
        )
    else:
        sleep = np.array([current["sleep_hours"]])

    screen = (
        create_range(
            max(0.0, current["screen_time"] - max_screen_reduction),
            current["screen_time"],
            0.5,
        )
        if "screen_time" in changeable_features
        else np.array([current["screen_time"]])
    )

    attendance = (
        create_range(
            current["attendance"],
            min(100.0, current["attendance"] + max_attendance_increase),
            1.0,
        )
        if "attendance" in changeable_features
        else np.array([current["attendance"]])
    )

    return {
        "study_hours": study,
        "attendance": attendance,
        "sleep_hours": sleep,
        "screen_time": screen,
    }


def generate_recommendation_candidates(
    current,
    target_score,
    changeable_features,
    max_study,
    min_sleep,
    max_screen_reduction,
    max_attendance_increase,
):
    values = generate_candidate_values(
        current,
        changeable_features,
        max_study,
        min_sleep,
        max_screen_reduction,
        max_attendance_increase,
    )
    current_prediction = predict_from_values(current)
    rows = []

    for study, attendance, sleep, screen in product(
        values["study_hours"],
        values["attendance"],
        values["sleep_hours"],
        values["screen_time"],
    ):
        candidate = current.copy()
        candidate.update(
            {
                "study_hours": float(study),
                "attendance": float(attendance),
                "sleep_hours": float(sleep),
                "screen_time": float(screen),
            }
        )

        total_hours = calculate_total_hours(candidate)
        if total_hours > 24:
            continue

        prediction = predict_from_values(candidate)
        cost = calculate_change_cost(current, candidate)
        difficulty = calculate_difficulty_score(cost)

        rows.append(
            {
                **candidate,
                "예상 환산 점수": prediction["display_score"],
                "원본 예측 점수": prediction["raw_score"],
                "현재 대비 점수 변화": prediction["display_score"] - current_prediction["display_score"],
                "변화 부담 수치": cost,
                "변화 부담": describe_change_cost(cost),
                "실행 난이도 점수": difficulty,
                "실행 난이도": describe_difficulty(difficulty),
                "목표 달성": prediction["display_score"] >= target_score,
                "현재와 다른 계획": has_any_change(current, candidate),
                "현재보다 개선": prediction["display_score"] > current_prediction["display_score"] + EPSILON,
                "주요 생활시간 합계": total_hours,
            }
        )

    return pd.DataFrame(rows)


def plan_key(row):
    return tuple(round(float(row[f]), 4) for f in FEATURE_COLS)


def choose_unique_row(dataframe, used_keys):
    for _, row in dataframe.iterrows():
        key = plan_key(row)
        if key not in used_keys:
            used_keys.add(key)
            return row
    return None


def select_recommendations(candidates, target_score):
    empty_result = {
        "target_reached": False,
        "recommendations": pd.DataFrame(),
        "candidate_pool_type": "후보 없음",
    }
    if candidates.empty:
        return empty_result

    changed = candidates[candidates["현재와 다른 계획"]].copy()
    if changed.empty:
        return {**empty_result, "candidate_pool_type": "변경 후보 없음"}

    reached = changed[changed["예상 환산 점수"] >= target_score].copy()
    target_reached = not reached.empty

    if target_reached:
        pool = reached
        pool_type = "목표 달성 후보"
        labels = ("최소 변화형", "균형형", "성과 우선형")
    else:
        improved = changed[changed["현재보다 개선"]].copy()
        pool = improved if not improved.empty else changed
        pool_type = "목표 미달 개선 후보" if not improved.empty else "목표 미달 변경 후보"
        labels = ("최소 변화 개선형", "균형 개선형", "최고 성과형")

    pool["균형 점수"] = (
        pool["예상 환산 점수"] / DISPLAY_SCORE_MAX * 0.65
        + (1.0 - np.clip(pool["변화 부담 수치"], 0.0, 1.0)) * 0.35
    )

    used = set()
    minimum = choose_unique_row(
        pool.sort_values(["변화 부담 수치", "예상 환산 점수"], ascending=[True, False]),
        used,
    )
    balanced = choose_unique_row(
        pool.sort_values(["균형 점수", "예상 환산 점수", "변화 부담 수치"], ascending=[False, False, True]),
        used,
    )
    performance = choose_unique_row(
        pool.sort_values(["예상 환산 점수", "변화 부담 수치"], ascending=[False, True]),
        used,
    )

    reasons = [
        "현재 생활과의 차이가 가장 작은 계획입니다.",
        "예상 성과와 생활 변화 부담을 함께 고려한 계획입니다.",
        "설정한 조건에서 가장 높은 학업성과 점수를 예측한 계획입니다.",
    ]

    rows = []
    for label, row, reason in zip(labels, [minimum, balanced, performance], reasons):
        if row is None:
            continue
        rows.append(
            {
                "추천 유형": label,
                "공부 시간": row["study_hours"],
                "출석률": row["attendance"],
                "수면 시간": row["sleep_hours"],
                "스크린타임": row["screen_time"],
                "신체활동": row["physical_activity"],
                "스트레스": row["stress"],
                "예상 환산 점수": row["예상 환산 점수"],
                "현재 대비 점수 변화": row["현재 대비 점수 변화"],
                "원본 예측 점수": row["원본 예측 점수"],
                "변화 부담": row["변화 부담"],
                "변화 부담 수치": row["변화 부담 수치"],
                "실행 난이도": row["실행 난이도"],
                "실행 난이도 점수": row["실행 난이도 점수"],
                "주요 생활시간 합계": row["주요 생활시간 합계"],
                "추천 이유": reason,
            }
        )

    return {
        "target_reached": target_reached,
        "recommendations": pd.DataFrame(rows),
        "candidate_pool_type": pool_type,
    }


# =========================================================
# 추천 설명 및 실행계획
# =========================================================
def build_change_summary(current, recommendation):
    mapping = {
        "study_hours": ("공부 시간", "공부 시간", "시간"),
        "attendance": ("출석률", "출석률", "%p"),
        "sleep_hours": ("수면 시간", "수면 시간", "시간"),
        "screen_time": ("스크린타임", "스크린타임", "시간"),
    }
    changes = []

    for current_key, (column, label, unit) in mapping.items():
        difference = float(recommendation[column]) - current[current_key]
        if abs(difference) < EPSILON:
            continue
        direction = "증가" if difference > 0 else "감소"
        digits = 0 if current_key == "attendance" else 1
        changes.append(f"{label} {abs(difference):.{digits}f}{unit} {direction}")

    return ", ".join(changes) if changes else "현재 생활 유지"


def generate_simple_action_plan(current, recommendation):
    study = float(recommendation["공부 시간"])
    attendance = float(recommendation["출석률"])
    sleep = float(recommendation["수면 시간"])
    screen = float(recommendation["스크린타임"])

    actions = [
        f"공부 {study:.1f}시간 {'확보' if study > current['study_hours'] else '유지'}",
        f"수면 {sleep:.1f}시간 {'확보' if sleep > current['sleep_hours'] else '유지'}",
        f"스크린타임 {screen:.1f}시간 이내",
    ]
    goals = [
        f"공부 {study:.1f}시간",
        f"수면 {sleep:.1f}시간",
    ]
    checks = ["공부 목표 달성 여부", "취침·기상 시간"]

    if attendance > current["attendance"]:
        actions.insert(1, f"출석률 {attendance:.0f}% 목표")
        goals.append(f"출석률 {attendance:.0f}%")
        checks.append("지각·결석 여부")

    if screen < current["screen_time"]:
        goals.append(f"스크린타임 {screen:.1f}시간 이하")
        checks.append("스크린타임 확인")

    weekend_actions = [
        "한 주 학습 내용 복습",
        f"수면 {sleep:.1f}시간 유지",
        f"스크린타임 {screen:.1f}시간 이하",
    ]

    weekly_rows = []
    for day in DAYS:
        weekend = day in ("토요일", "일요일")
        weekly_rows.append(
            {
                "요일": day,
                "공부 목표": f"복습 포함 {study:.1f}시간 이내" if weekend else f"{study:.1f}시간",
                "출석 목표": "해당 없음" if weekend else f"{attendance:.0f}% 수준 유지",
                "수면 목표": f"{sleep:.1f}시간",
                "스크린타임 목표": f"{screen:.1f}시간 이하",
            }
        )

    return {
        "weekday_actions": actions,
        "weekend_actions": weekend_actions,
        "core_goals": goals,
        "check_items": checks,
        "weekly_df": pd.DataFrame(weekly_rows),
    }


def generate_three_week_plan(current, recommendation):
    targets = {
        "study_hours": float(recommendation["공부 시간"]),
        "attendance": float(recommendation["출석률"]),
        "sleep_hours": float(recommendation["수면 시간"]),
        "screen_time": float(recommendation["스크린타임"]),
    }
    settings = [
        ("1주차", 1 / 3, "생활 변화에 적응"),
        ("2주차", 2 / 3, "목표 수준에 근접"),
        ("3주차", 1.0, "최종 추천 목표 실천"),
    ]

    rows = []
    for week, ratio, goal in settings:
        rows.append(
            {
                "주차": week,
                "공부 시간": round(current["study_hours"] + (targets["study_hours"] - current["study_hours"]) * ratio, 1),
                "출석률": round(current["attendance"] + (targets["attendance"] - current["attendance"]) * ratio, 1),
                "수면 시간": round(current["sleep_hours"] + (targets["sleep_hours"] - current["sleep_hours"]) * ratio, 1),
                "스크린타임": round(current["screen_time"] + (targets["screen_time"] - current["screen_time"]) * ratio, 1),
                "핵심 목표": goal,
            }
        )
    return pd.DataFrame(rows)


def make_recommendation_comparison(current, recommendations):
    rows = []
    for _, row in recommendations.iterrows():
        rows.append(
            {
                "추천 유형": row["추천 유형"],
                "공부 시간 변화": row["공부 시간"] - current["study_hours"],
                "출석률 변화": row["출석률"] - current["attendance"],
                "수면 시간 변화": row["수면 시간"] - current["sleep_hours"],
                "스크린타임 변화": row["스크린타임"] - current["screen_time"],
                "예상 환산 점수": row["예상 환산 점수"],
                "현재 대비 점수 변화": row["현재 대비 점수 변화"],
                "변화 부담": row["변화 부담"],
                "실행 난이도": row["실행 난이도"],
            }
        )
    return pd.DataFrame(rows)


def build_recommendation_card_html(row, current, selected_type):
    selected = row["추천 유형"] == selected_type
    card_class = "recommendation-card-selected" if selected else "recommendation-card"
    badge = '<div class="selected-badge">선택된 추천안</div>' if selected else ""
    return (
        f'<div class="{card_class}">{badge}'
        f'<div class="card-title">{escape(str(row["추천 유형"]))}</div>'
        '<div class="card-score-label">예상 학업성과 점수</div>'
        f'<div class="card-score">{row["예상 환산 점수"]:.2f} / {DISPLAY_SCORE_MAX:.2f}</div>'
        f'<div class="card-delta-positive">↑ {row["현재 대비 점수 변화"]:+.2f}</div>'
        f'<p><b>공부 시간:</b> {row["공부 시간"]:.1f}시간</p>'
        f'<p><b>출석률:</b> {row["출석률"]:.0f}%</p>'
        f'<p><b>수면 시간:</b> {row["수면 시간"]:.1f}시간</p>'
        f'<p><b>스크린타임:</b> {row["스크린타임"]:.1f}시간</p>'
        f'<p><b>실행 난이도:</b> {escape(str(row["실행 난이도"]))} ({int(row["실행 난이도 점수"])}/5)</p>'
        f'<div class="card-reason">{escape(str(row["추천 이유"]))}</div>'
        f'<small>{escape(build_change_summary(current, row))}</small></div>'
    )


# =========================================================
# 엑셀 스타일 및 유틸리티
# =========================================================
TITLE_FILL = PatternFill("solid", fgColor="1F4E78")
SUBTITLE_FILL = PatternFill("solid", fgColor="D9EAF7")
HEADER_FILL = PatternFill("solid", fgColor="5B9BD5")
LIGHT_BLUE_FILL = PatternFill("solid", fgColor="EAF3F8")
GREEN_FILL = PatternFill("solid", fgColor="E2F0D9")
YELLOW_FILL = PatternFill("solid", fgColor="FFF2CC")
RED_FILL = PatternFill("solid", fgColor="FCE4D6")
GRAY_FILL = PatternFill("solid", fgColor="F2F2F2")
THIN = Side(style="thin", color="D9E1F2")
CELL_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def set_excel_title(ws, title, subtitle, last_column):
    end = get_column_letter(last_column)
    ws.merge_cells(f"A1:{end}1")
    ws["A1"] = title
    ws["A1"].fill = TITLE_FILL
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=18)
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells(f"A2:{end}2")
    ws["A2"] = subtitle
    ws["A2"].fill = SUBTITLE_FILL
    ws["A2"].font = Font(bold=True, size=11)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 35

    ws.merge_cells(f"A3:{end}3")
    ws["A3"] = f"생성일: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A3"].font = Font(italic=True, color="666666", size=9)
    ws["A3"].alignment = Alignment(horizontal="right")


def set_excel_sheet_options(ws, landscape=True, freeze_panes="A5"):
    ws.freeze_panes = freeze_panes
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5


def write_excel_table(ws, start_row, headers, rows, widths=None, number_formats=None):
    for column, header in enumerate(headers, 1):
        cell = ws.cell(start_row, column, header)
        cell.fill = HEADER_FILL
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = CELL_BORDER

    ws.row_dimensions[start_row].height = 28

    for row_index, values in enumerate(rows, start_row + 1):
        for column, value in enumerate(values, 1):
            cell = ws.cell(row_index, column, value)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = CELL_BORDER
            if row_index % 2 == 0:
                cell.fill = LIGHT_BLUE_FILL
            if number_formats and column in number_formats and not isinstance(value, str):
                cell.number_format = number_formats[column]
        ws.row_dimensions[row_index].height = 25

    if widths:
        for column, width in widths.items():
            ws.column_dimensions[get_column_letter(column)].width = width

    if rows:
        ws.auto_filter.ref = f"A{start_row}:{get_column_letter(len(headers))}{start_row + len(rows)}"


def write_key_value_section(ws, start_row, title, values, value_width=38):
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=2)
    title_cell = ws.cell(start_row, 1, title)
    title_cell.fill = SUBTITLE_FILL
    title_cell.font = Font(bold=True, size=13, color="1F4E78")
    title_cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[start_row].height = 25

    row = start_row + 1
    for label, value in values:
        label_cell = ws.cell(row, 1, label)
        value_cell = ws.cell(row, 2, value)
        label_cell.fill = GRAY_FILL
        label_cell.font = Font(bold=True)
        for cell in (label_cell, value_cell):
            cell.border = CELL_BORDER
            cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[row].height = 24
        row += 1

    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = value_width
    return row


def difficulty_fill(score):
    if score <= 2:
        return GREEN_FILL
    if score == 3:
        return YELLOW_FILL
    return RED_FILL


def format_excel_change(value):
    return "유지" if abs(float(value)) < EPSILON else float(value)


# =========================================================
# 전체 추천안 엑셀
# =========================================================
def recommendations_to_xlsx(recommendations, comparison_df, current, target_score):
    wb = Workbook()
    ws = wb.active
    ws.title = "추천안 비교"

    set_excel_title(
        ws,
        "Campus Twin 전체 추천안 비교",
        "목표 점수와 생활 제약조건 안에서 생성된 추천안을 핵심 항목 중심으로 비교합니다.",
        10,
    )

    headers = [
        "추천 유형",
        "공부 시간",
        "출석률",
        "수면 시간",
        "스크린타임",
        "예상 학업성과 점수",
        "점수 변화",
        "변화 부담",
        "실행 난이도",
        "추천 이유",
    ]
    rows = [
        [
            row["추천 유형"],
            float(row["공부 시간"]),
            float(row["출석률"]),
            float(row["수면 시간"]),
            float(row["스크린타임"]),
            float(row["예상 환산 점수"]),
            float(row["현재 대비 점수 변화"]),
            row["변화 부담"],
            f"{row['실행 난이도']} ({int(row['실행 난이도 점수'])}/5)",
            row["추천 이유"],
        ]
        for _, row in recommendations.iterrows()
    ]

    write_excel_table(
        ws,
        5,
        headers,
        rows,
        widths={1: 20, 2: 12, 3: 12, 4: 12, 5: 14, 6: 20, 7: 13, 8: 12, 9: 18, 10: 42},
        number_formats={2: "0.0", 3: "0.0", 4: "0.0", 5: "0.0", 6: "0.00", 7: "+0.00;-0.00;0.00"},
    )

    for row_index, (_, row) in enumerate(recommendations.iterrows(), start=6):
        ws.cell(row_index, 9).fill = difficulty_fill(int(row["실행 난이도 점수"]))

    set_excel_sheet_options(ws, True, "A6")

    change_ws = wb.create_sheet("변화 비교")
    set_excel_title(
        change_ws,
        "현재 생활 대비 변화 비교",
        "각 추천안이 현재 생활습관에서 얼마나 달라지는지 보여줍니다.",
        len(comparison_df.columns),
    )
    write_excel_table(
        change_ws,
        5,
        list(comparison_df.columns),
        comparison_df.astype(object).values.tolist(),
        widths={1: 21, 2: 17, 3: 17, 4: 17, 5: 19, 6: 15, 7: 18, 8: 13, 9: 15},
        number_formats={2: "+0.00;-0.00;0.00", 3: "+0.00;-0.00;0.00", 4: "+0.00;-0.00;0.00", 5: "+0.00;-0.00;0.00", 6: "0.00", 7: "+0.00;-0.00;0.00"},
    )
    set_excel_sheet_options(change_ws, True, "A6")

    condition_ws = wb.create_sheet("입력 조건")
    set_excel_title(
        condition_ws,
        "Campus Twin 입력 조건",
        "추천 계획을 생성할 때 사용한 현재 생활습관과 목표 점수입니다.",
        2,
    )
    current_prediction = predict_from_values(current)
    write_key_value_section(
        condition_ws,
        5,
        "현재 상태 및 목표",
        [
            ("현재 공부 시간", f"{current['study_hours']:.1f}시간"),
            ("현재 출석률", f"{current['attendance']:.1f}%"),
            ("현재 수면 시간", f"{current['sleep_hours']:.1f}시간"),
            ("현재 스크린타임", f"{current['screen_time']:.1f}시간"),
            ("현재 신체활동", f"{current['physical_activity']:.1f}시간"),
            ("현재 스트레스", f"{current['stress']:.1f}단계"),
            ("현재 학업성과 점수", f"{current_prediction['display_score']:.2f} / 4.50"),
            ("목표 학업성과 점수", f"{target_score:.2f} / 4.50"),
        ],
        30,
    )
    set_excel_sheet_options(condition_ws, False, "A6")

    detail_ws = wb.create_sheet("세부 데이터")
    set_excel_title(
        detail_ws,
        "추천안 세부 계산 데이터",
        "모델 원본 예측값과 변화 부담 수치 등 분석용 데이터를 포함합니다.",
        len(recommendations.columns),
    )
    widths = {i: 16 for i in range(1, len(recommendations.columns) + 1)}
    if "추천 이유" in recommendations.columns:
        widths[list(recommendations.columns).index("추천 이유") + 1] = 45
    write_excel_table(
        detail_ws,
        5,
        list(recommendations.columns),
        recommendations.astype(object).values.tolist(),
        widths=widths,
    )
    set_excel_sheet_options(detail_ws, True, "A6")

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


# =========================================================
# 선택한 계획 엑셀
# =========================================================
def selected_plan_to_xlsx(selected_plan, weekly_plan, three_week_plan, current, completed_days):
    wb = Workbook()
    ws = wb.active
    ws.title = "계획 요약"

    set_excel_title(
        ws,
        "Campus Twin 선택 계획 요약",
        "선택한 추천안의 핵심 정보와 현재 생활 대비 변화를 정리한 문서입니다.",
        4,
    )

    difficulty_score = int(selected_plan["실행 난이도 점수"])
    progress = int(round(len(completed_days) / len(DAYS) * 100))

    next_row = write_key_value_section(
        ws,
        5,
        "선택한 추천안",
        [
            ("추천 유형", selected_plan["추천 유형"]),
            ("예상 학업성과 점수", f"{selected_plan['예상 환산 점수']:.2f} / 4.50"),
            ("현재 대비 점수 변화", f"{selected_plan['현재 대비 점수 변화']:+.2f}"),
            ("실행 난이도", f"{selected_plan['실행 난이도']} ({difficulty_score}/5)"),
            ("핵심 생활 변화", build_change_summary(current, selected_plan)),
            ("추천 이유", selected_plan["추천 이유"]),
            ("이번 주 진행률", f"{len(completed_days)}일 / 7일 ({progress}%)"),
        ],
        52,
    )

    ws[f"B{next_row - 4}"].fill = difficulty_fill(difficulty_score)

    current_target_rows = [
        [
            "공부 시간",
            current["study_hours"],
            float(selected_plan["공부 시간"]),
            format_excel_change(float(selected_plan["공부 시간"]) - current["study_hours"]),
        ],
        [
            "출석률",
            current["attendance"],
            float(selected_plan["출석률"]),
            format_excel_change(float(selected_plan["출석률"]) - current["attendance"]),
        ],
        [
            "수면 시간",
            current["sleep_hours"],
            float(selected_plan["수면 시간"]),
            format_excel_change(float(selected_plan["수면 시간"]) - current["sleep_hours"]),
        ],
        [
            "스크린타임",
            current["screen_time"],
            float(selected_plan["스크린타임"]),
            format_excel_change(float(selected_plan["스크린타임"]) - current["screen_time"]),
        ],
    ]

    table_start = next_row + 2
    write_excel_table(
        ws,
        table_start,
        ["생활요인", "현재", "추천 목표", "변화"],
        current_target_rows,
        widths={1: 18, 2: 14, 3: 14, 4: 14},
        number_formats={2: "0.0", 3: "0.0", 4: "+0.0;-0.0;0.0"},
    )

    for offset, row_values in enumerate(current_target_rows, 1):
        excel_row = table_start + offset
        if row_values[3] == "유지":
            for column in range(1, 5):
                cell = ws.cell(excel_row, column)
                cell.fill = GRAY_FILL
                cell.font = Font(color="666666")
            ws.cell(excel_row, 4).font = Font(bold=True, color="666666")

    set_excel_sheet_options(ws, False, "A5")

    three_ws = wb.create_sheet("3주 실행계획")
    set_excel_title(
        three_ws,
        "3주 단계적 실행계획",
        "생활 변화 부담을 줄이기 위해 목표 수준에 단계적으로 접근합니다.",
        6,
    )
    three_rows = three_week_plan[
        ["주차", "공부 시간", "출석률", "수면 시간", "스크린타임", "핵심 목표"]
    ].astype(object).values.tolist()
    write_excel_table(
        three_ws,
        5,
        ["주차", "공부 시간", "출석률", "수면 시간", "스크린타임", "핵심 목표"],
        three_rows,
        widths={1: 13, 2: 14, 3: 14, 4: 14, 5: 16, 6: 30},
        number_formats={2: "0.0", 3: "0.0", 4: "0.0", 5: "0.0"},
    )
    for row_index, fill in zip(range(6, 9), [GREEN_FILL, YELLOW_FILL, RED_FILL]):
        for column in range(1, 7):
            three_ws.cell(row_index, column).fill = fill
    set_excel_sheet_options(three_ws, True, "A6")

    weekly_ws = wb.create_sheet("주간 실천표")
    set_excel_title(
        weekly_ws,
        "요일별 주간 실천표",
        "요일별 목표를 확인하고 실천 후 완료 여부를 직접 표시할 수 있습니다.",
        6,
    )
    weekly_rows = []
    for _, row in weekly_plan["weekly_df"].iterrows():
        day = row["요일"]
        weekly_rows.append(
            [
                day,
                row["공부 목표"],
                row["출석 목표"],
                row["수면 목표"],
                row["스크린타임 목표"],
                "완료" if day in completed_days else "미완료",
            ]
        )
    write_excel_table(
        weekly_ws,
        5,
        ["요일", "공부 목표", "출석 목표", "수면 목표", "스크린타임 목표", "완료 여부"],
        weekly_rows,
        widths={1: 12, 2: 24, 3: 20, 4: 16, 5: 22, 6: 14},
    )
    for row_index, row_values in enumerate(weekly_rows, 6):
        status = weekly_ws.cell(row_index, 6)
        if row_values[5] == "완료":
            status.fill = GREEN_FILL
            status.font = Font(bold=True, color="008000")
        else:
            status.fill = GRAY_FILL
            status.font = Font(color="666666")
    set_excel_sheet_options(weekly_ws, True, "A6")

    checklist_ws = wb.create_sheet("실천 체크리스트")
    set_excel_title(
        checklist_ws,
        "일주일 실천 체크리스트",
        "평일·주말 핵심 행동과 매일 확인할 항목을 정리했습니다.",
        3,
    )
    checklist_rows = (
        [["평일 핵심 행동", action, "□"] for action in weekly_plan["weekday_actions"]]
        + [["주말 핵심 행동", action, "□"] for action in weekly_plan["weekend_actions"]]
        + [["이번 주 목표", goal, "□"] for goal in weekly_plan["core_goals"]]
        + [["매일 확인", item, "□"] for item in weekly_plan["check_items"]]
    )
    write_excel_table(
        checklist_ws,
        5,
        ["구분", "실천 내용", "확인"],
        checklist_rows,
        widths={1: 20, 2: 52, 3: 12},
    )
    set_excel_sheet_options(checklist_ws, False, "A6")

    detail_ws = wb.create_sheet("세부 분석")
    set_excel_title(
        detail_ws,
        "선택 계획 세부 분석",
        "모델 예측값과 변화 부담 등 분석용 세부 정보를 포함합니다.",
        2,
    )
    write_key_value_section(
        detail_ws,
        5,
        "모델 및 추천 세부 정보",
        [
            ("추천 유형", selected_plan["추천 유형"]),
            ("공부 시간", float(selected_plan["공부 시간"])),
            ("출석률", float(selected_plan["출석률"])),
            ("수면 시간", float(selected_plan["수면 시간"])),
            ("스크린타임", float(selected_plan["스크린타임"])),
            ("신체활동", float(selected_plan["신체활동"])),
            ("스트레스", float(selected_plan["스트레스"])),
            ("예상 학업성과 점수", float(selected_plan["예상 환산 점수"])),
            ("원본 예측 점수", float(selected_plan["원본 예측 점수"])),
            ("현재 대비 점수 변화", float(selected_plan["현재 대비 점수 변화"])),
            ("변화 부담", selected_plan["변화 부담"]),
            ("변화 부담 수치", float(selected_plan["변화 부담 수치"])),
            ("실행 난이도", selected_plan["실행 난이도"]),
            ("실행 난이도 점수", int(selected_plan["실행 난이도 점수"])),
            ("주요 생활시간 합계", float(selected_plan["주요 생활시간 합계"])),
            ("추천 이유", selected_plan["추천 이유"]),
        ],
        46,
    )
    set_excel_sheet_options(detail_ws, False, "A6")

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


# =========================================================
# 세션 상태
# =========================================================
def clear_daily_checkboxes():
    for key in list(st.session_state.keys()):
        if str(key).startswith("daily_check_"):
            del st.session_state[key]


DEFAULT_SESSION_VALUES = {
    "current_values": None,
    "prediction_result": None,
    "sensitivity_df": None,
    "comparison_result": None,
    "recommendation_result": None,
    "recommendation_conditions": None,
    "selected_recommendation_type": None,
    "previous_recommendation_type": None,
}
for key, value in DEFAULT_SESSION_VALUES.items():
    st.session_state.setdefault(key, value)


# =========================================================
# 상단
# =========================================================
st.title("🎓 Campus Twin")
st.subheader("AI 기반 대학생활 시뮬레이션 및 생활계획 추천")
st.write(
    """
    생활습관을 바탕으로 학업성과를 분석하고 미래 계획을 시뮬레이션합니다.
    목표와 현실적인 조건을 입력하면 추천안을 탐색하고,
    선택한 추천안을 3주 단계계획과 일주일 실천표로 변환합니다.
    """
)
st.warning(
    "학업성과 환산 점수는 실제 학교 성적이나 공식 GPA가 아닙니다. "
    "모델의 원본 예측값을 4.5점 범위로 변환한 참고용 지표입니다."
)

tab1, tab2, tab3, tab4, tab5 = st.tabs(
    ["현재 상태", "개인별 분석", "미래 시뮬레이션", "추천 계획", "프로젝트 설명"]
)


# =========================================================
# 현재 상태
# =========================================================
with tab1:
    st.header("현재 생활습관 입력")
    left, right = st.columns(2)

    with left:
        study_hours = st.number_input("하루 공부 시간", 0.0, 12.0, 4.0, 0.5, key="current_study")
        attendance = st.number_input("출석률", 0.0, 100.0, 85.0, 1.0, key="current_attendance")
        sleep_hours = st.number_input("하루 수면 시간", 0.0, 12.0, 7.0, 0.5, key="current_sleep")

    with right:
        screen_time = st.number_input("하루 스크린타임", 0.0, 16.0, 5.0, 0.5, key="current_screen")
        physical_activity = st.number_input("하루 신체활동 시간", 0.0, 8.0, 1.0, 0.5, key="current_activity")
        stress = st.slider("스트레스 수준", 1.0, 10.0, 5.0, 0.5, key="current_stress")

    current_values = {
        "study_hours": study_hours,
        "attendance": attendance,
        "sleep_hours": sleep_hours,
        "screen_time": screen_time,
        "physical_activity": physical_activity,
        "stress": stress,
    }

    total_hours = calculate_total_hours(current_values)
    st.caption(f"입력된 주요 생활시간 합계: {total_hours:.1f}시간")
    if total_hours > 24:
        st.warning("주요 생활시간 합이 하루 24시간을 초과합니다.")

    if st.button("현재 상태 분석하기", type="primary"):
        try:
            st.session_state.current_values = current_values.copy()
            st.session_state.prediction_result = predict_from_values(current_values)
            st.session_state.sensitivity_df = analyze_sensitivity(current_values)
            st.session_state.comparison_result = None
            st.session_state.recommendation_result = None
            st.session_state.recommendation_conditions = None
            st.session_state.selected_recommendation_type = None
            st.session_state.previous_recommendation_type = None
            clear_daily_checkboxes()
            st.session_state.pop("recommendation_selector", None)
        except Exception as error:
            st.error("예측 중 오류가 발생했습니다.")
            st.exception(error)

    if st.session_state.prediction_result is not None:
        result = st.session_state.prediction_result
        st.divider()
        st.subheader("예측 결과")
        col1, col2 = st.columns(2)
        col1.metric("학업성과 환산 점수", f"{result['display_score']:.2f} / {DISPLAY_SCORE_MAX:.2f}")
        col2.metric("모델 원본 예측 점수", f"{result['raw_score']:.3f}")
        st.caption("예측 결과는 실제 성적을 확정하거나 보장하지 않습니다.")


# =========================================================
# 개인별 분석
# =========================================================
with tab2:
    st.header("개인별 민감도 분석")

    if st.session_state.current_values is None or st.session_state.sensitivity_df is None:
        st.info("'현재 상태' 탭에서 먼저 분석해 주세요.")
    else:
        sensitivity = st.session_state.sensitivity_df.copy()
        display = sensitivity.copy()
        display["현재 값"] = display["현재 값"].map(lambda x: f"{x:.1f}")
        display["변경 값"] = display["변경 값"].map(lambda x: f"{x:.1f}")
        display["변경 후 환산 점수"] = display["변경 후 환산 점수"].map(lambda x: f"{x:.2f}")
        display["환산 점수 변화량"] = display["환산 점수 변화량"].map(lambda x: f"{x:+.3f}")
        display["원본 점수 변화량"] = display["원본 점수 변화량"].map(lambda x: f"{x:+.4f}")

        st.dataframe(display, use_container_width=True, hide_index=True)
        chart = sensitivity[["생활요인 변화", "환산 점수 변화량"]].set_index("생활요인 변화")
        st.bar_chart(chart, use_container_width=True)

        best = sensitivity.iloc[0]
        st.success(
            f"가장 큰 예측 상승 요인: {best['생활요인 변화']} "
            f"({best['환산 점수 변화량']:+.3f})"
        )
        st.caption("모델 예측 반응이며 실제 인과관계를 의미하지 않습니다.")


# =========================================================
# 미래 시뮬레이션
# =========================================================
with tab3:
    st.header("미래 생활계획 시뮬레이션")

    if st.session_state.current_values is None:
        st.info("'현재 상태' 탭에서 먼저 분석해 주세요.")
    else:
        current = st.session_state.current_values
        left, right = st.columns(2)

        with left:
            future_study = st.number_input("계획 공부 시간", 0.0, 12.0, float(current["study_hours"]), 0.5, key="future_study")
            future_attendance = st.number_input("계획 출석률", 0.0, 100.0, float(current["attendance"]), 1.0, key="future_attendance")
            future_sleep = st.number_input("계획 수면 시간", 0.0, 12.0, float(current["sleep_hours"]), 0.5, key="future_sleep")

        with right:
            future_screen = st.number_input("계획 스크린타임", 0.0, 16.0, float(current["screen_time"]), 0.5, key="future_screen")
            future_activity = st.number_input("계획 신체활동 시간", 0.0, 8.0, float(current["physical_activity"]), 0.5, key="future_activity")
            future_stress = st.slider("계획 스트레스 수준", 1.0, 10.0, float(current["stress"]), 0.5, key="future_stress")

        future = {
            "study_hours": future_study,
            "attendance": future_attendance,
            "sleep_hours": future_sleep,
            "screen_time": future_screen,
            "physical_activity": future_activity,
            "stress": future_stress,
        }

        if st.button("현재 계획과 비교하기", type="primary"):
            try:
                st.session_state.comparison_result = compare_plans(current, future)
            except Exception as error:
                st.error("미래 계획 비교 중 오류가 발생했습니다.")
                st.exception(error)

        if st.session_state.comparison_result is not None:
            result = st.session_state.comparison_result
            st.divider()
            col1, col2, col3 = st.columns(3)
            col1.metric("현재 학업성과 점수", f"{result['current_display_score']:.2f}")
            col2.metric("미래 학업성과 점수", f"{result['future_display_score']:.2f}")
            col3.metric("예상 변화", f"{result['display_score_change']:+.2f}")
            st.dataframe(result["comparison_df"], use_container_width=True, hide_index=True)


# =========================================================
# 추천 계획
# =========================================================
with tab4:
    st.header("목표 기반 추천 계획")

    if st.session_state.current_values is None:
        st.info("'현재 상태' 탭에서 먼저 분석해 주세요.")
    else:
        current = st.session_state.current_values
        current_prediction = predict_from_values(current)
        st.metric(
            "현재 학업성과 환산 점수",
            f"{current_prediction['display_score']:.2f} / {DISPLAY_SCORE_MAX:.2f}",
        )

        st.subheader("목표 및 현실성 조건")
        left, right = st.columns(2)

        with left:
            default_target = min(
                DISPLAY_SCORE_MAX,
                round(current_prediction["display_score"] + 0.3, 1),
            )
            target_score = st.number_input(
                "목표 학업성과 환산 점수",
                0.0,
                DISPLAY_SCORE_MAX,
                float(default_target),
                0.1,
            )
            max_study = st.number_input(
                "하루 최대 공부 시간",
                float(current["study_hours"]),
                12.0,
                float(min(12.0, current["study_hours"] + 2.0)),
                0.5,
            )
            min_sleep = st.number_input(
                "반드시 확보할 최소 수면 시간",
                0.0,
                12.0,
                float(current["sleep_hours"]),
                0.5,
            )

        with right:
            max_screen_reduction = st.number_input(
                "줄일 수 있는 최대 스크린타임",
                0.0,
                float(current["screen_time"]),
                float(min(2.0, current["screen_time"])),
                0.5,
            )
            max_attendance_increase = st.number_input(
                "가능한 출석률 최대 증가폭",
                0.0,
                float(100.0 - current["attendance"]),
                float(min(10.0, 100.0 - current["attendance"])),
                1.0,
            )
            changeable_labels = st.multiselect(
                "변경 가능한 생활요인",
                ["공부 시간", "출석률", "수면 시간", "스크린타임"],
                ["공부 시간", "출석률", "수면 시간", "스크린타임"],
            )

        changeable_features = [LABEL_TO_FEATURE[label] for label in changeable_labels]

        if st.button("추천 계획 찾기", type="primary"):
            if not changeable_features:
                st.warning("변경 가능한 생활요인을 한 개 이상 선택해 주세요.")
            elif min_sleep > current["sleep_hours"] + 1.5:
                st.warning("최소 수면시간을 현재보다 1.5시간 이내로 설정해 주세요.")
            else:
                try:
                    with st.spinner("가능한 생활계획을 탐색하고 있습니다."):
                        candidates = generate_recommendation_candidates(
                            current,
                            target_score,
                            changeable_features,
                            max_study,
                            min_sleep,
                            max_screen_reduction,
                            max_attendance_increase,
                        )
                        st.session_state.recommendation_result = select_recommendations(
                            candidates, target_score
                        )
                        st.session_state.recommendation_conditions = {
                            "target_score": target_score,
                            "candidate_count": len(candidates),
                        }
                        st.session_state.selected_recommendation_type = None
                        st.session_state.previous_recommendation_type = None
                        clear_daily_checkboxes()
                        st.session_state.pop("recommendation_selector", None)
                except Exception as error:
                    st.error("추천 계획 탐색 중 오류가 발생했습니다.")
                    st.exception(error)

        if st.session_state.recommendation_result is not None:
            result = st.session_state.recommendation_result
            conditions = st.session_state.recommendation_conditions
            recommendations = result["recommendations"]

            st.divider()
            st.header("추천 결과")
            st.caption(f"총 {conditions['candidate_count']:,}개 후보 계획을 평가했습니다.")

            if recommendations.empty:
                st.error("추천안을 생성하지 못했습니다. 조건 범위를 넓혀 주세요.")
            else:
                if result["target_reached"]:
                    st.success("목표 점수를 달성하는 후보를 찾았습니다.")
                else:
                    st.warning("목표에는 도달하지 못했지만 현재보다 개선되는 후보를 제시합니다.")

                types = recommendations["추천 유형"].tolist()
                default_index = (
                    types.index(st.session_state.selected_recommendation_type)
                    if st.session_state.selected_recommendation_type in types
                    else 0
                )
                selected_type = st.radio(
                    "일주일 계획으로 사용할 추천안을 선택하세요.",
                    types,
                    index=default_index,
                    horizontal=True,
                    key="recommendation_selector",
                )

                if st.session_state.previous_recommendation_type != selected_type:
                    clear_daily_checkboxes()
                    st.session_state.previous_recommendation_type = selected_type
                st.session_state.selected_recommendation_type = selected_type

                st.subheader("추천안 카드")
                columns = st.columns(len(recommendations))
                for column, (_, row) in zip(columns, recommendations.iterrows()):
                    with column:
                        st.markdown(
                            build_recommendation_card_html(row, current, selected_type),
                            unsafe_allow_html=True,
                        )

                selected_plan = recommendations[
                    recommendations["추천 유형"] == selected_type
                ].iloc[0]

                st.success(f"선택한 추천안: **{selected_type}**")
                st.header("선택한 계획 요약")
                col1, col2, col3 = st.columns(3)
                col1.metric(
                    "예상 학업성과 점수",
                    f"{selected_plan['예상 환산 점수']:.2f} / {DISPLAY_SCORE_MAX:.2f}",
                )
                col2.metric(
                    "현재 대비 변화",
                    f"{selected_plan['현재 대비 점수 변화']:+.2f}",
                )
                col3.metric(
                    "실행 난이도",
                    f"{selected_plan['실행 난이도 점수']} / 5",
                )

                st.write("**생활 변화:** " + build_change_summary(current, selected_plan))
                st.write("**추천 이유:** " + selected_plan["추천 이유"])

                css_class, title, message = get_difficulty_message(
                    int(selected_plan["실행 난이도 점수"])
                )
                st.markdown(
                    f'<div class="{css_class}"><b>{escape(title)}</b><br>{escape(message)}</div>',
                    unsafe_allow_html=True,
                )

                st.divider()
                st.header("3주 단계적 실행계획")
                three_week_plan = generate_three_week_plan(current, selected_plan)
                week_columns = st.columns(3)

                for column, (_, week) in zip(week_columns, three_week_plan.iterrows()):
                    with column:
                        st.markdown(
                            (
                                '<div class="step-card">'
                                f'<h3>{escape(str(week["주차"]))}</h3>'
                                f'<p><b>공부:</b> {week["공부 시간"]:.1f}시간</p>'
                                f'<p><b>출석률:</b> {week["출석률"]:.1f}%</p>'
                                f'<p><b>수면:</b> {week["수면 시간"]:.1f}시간</p>'
                                f'<p><b>스크린타임:</b> {week["스크린타임"]:.1f}시간</p>'
                                f'<p><b>핵심:</b> {escape(str(week["핵심 목표"]))}</p>'
                                "</div>"
                            ),
                            unsafe_allow_html=True,
                        )

                weekly_plan = generate_simple_action_plan(current, selected_plan)
                st.divider()
                st.header("일주일 실천계획")

                weekday_col, weekend_col = st.columns(2)
                with weekday_col:
                    st.subheader("평일 핵심 행동")
                    for action in weekly_plan["weekday_actions"]:
                        st.write(f"- {action}")
                with weekend_col:
                    st.subheader("주말 핵심 행동")
                    for action in weekly_plan["weekend_actions"]:
                        st.write(f"- {action}")

                goal_col, check_col = st.columns(2)
                with goal_col:
                    st.subheader("이번 주 목표")
                    for goal in weekly_plan["core_goals"]:
                        st.write(f"✅ {goal}")
                with check_col:
                    st.subheader("매일 확인")
                    for item in weekly_plan["check_items"]:
                        st.write(f"□ {item}")

                st.subheader("요일별 실천표")
                st.dataframe(
                    weekly_plan["weekly_df"],
                    use_container_width=True,
                    hide_index=True,
                )

                st.divider()
                st.header("이번 주 실천 현황")
                st.write("실천을 완료한 요일을 체크하세요.")

                day_columns = st.columns(7)
                completed_days = []
                for column, day in zip(day_columns, DAYS):
                    with column:
                        checked = st.checkbox(
                            day.replace("요일", ""),
                            key=f"daily_check_{selected_type}_{day}",
                        )
                    if checked:
                        completed_days.append(day)

                ratio = len(completed_days) / len(DAYS)
                percent = int(round(ratio * 100))
                st.progress(ratio)
                st.markdown(
                    f'<div class="progress-summary">{len(completed_days)}일 / 7일 완료 · 진행률 {percent}%</div>',
                    unsafe_allow_html=True,
                )

                if len(completed_days) == 0:
                    st.info("첫 실천일을 완료하면 체크해 보세요.")
                elif len(completed_days) < 4:
                    st.info("좋은 시작입니다. 작은 실천을 계속 이어가세요.")
                elif len(completed_days) < 7:
                    st.success("절반 이상 완료했습니다. 목표까지 조금 남았습니다.")
                else:
                    st.success("이번 주 계획을 모두 완료했습니다! 🎉")

                if st.button("이번 주 체크 초기화"):
                    clear_daily_checkboxes()
                    clear_completed_days ()
                    st.rerun()

                st.divider()
                comparison = make_recommendation_comparison(current, recommendations)

                with st.expander("전체 추천 계획 비교 보기"):
                    st.dataframe(
                        recommendations[
                            [
                                "추천 유형",
                                "공부 시간",
                                "출석률",
                                "수면 시간",
                                "스크린타임",
                                "예상 환산 점수",
                                "현재 대비 점수 변화",
                                "변화 부담",
                                "실행 난이도",
                                "추천 이유",
                            ]
                        ],
                        use_container_width=True,
                        hide_index=True,
                    )
                    st.dataframe(comparison, use_container_width=True, hide_index=True)

                all_xlsx = recommendations_to_xlsx(
                    recommendations,
                    comparison,
                    current,
                    conditions["target_score"],
                )
                selected_xlsx = selected_plan_to_xlsx(
                    selected_plan,
                    weekly_plan,
                    three_week_plan,
                    current,
                    completed_days,
                )

                col1, col2 = st.columns(2)
                with col1:
                    st.download_button(
                        "전체 추천안 엑셀 다운로드",
                        all_xlsx,
                        "campus_twin_all_recommendations.xlsx",
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                with col2:
                    st.download_button(
                        "선택한 계획 엑셀 다운로드",
                        selected_xlsx,
                        "campus_twin_selected_plan.xlsx",
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )

                st.caption(
                    "선택한 계획 엑셀에는 현재 체크한 요일과 진행률이 함께 저장됩니다."
                )


# =========================================================
# 프로젝트 설명
# =========================================================
with tab5:
    st.header("프로젝트 설명")
    st.subheader("프로젝트 목표")
    st.write(
        """
        Campus Twin은 대학생의 생활습관을 입력받아 학업성과를 예측하고,
        미래 생활계획과 목표 기반 추천안을 제공하는 웹서비스입니다.
        사용자는 추천안을 선택하고 3주 단계계획과 일주일 실천표를 확인할 수 있습니다.
        """
    )

    st.subheader("사용 변수")
    st.dataframe(
        pd.DataFrame(
            {
                "변수명": FEATURE_COLS,
                "설명": [FEATURE_LABELS[column] for column in FEATURE_COLS],
            }
        ),
        use_container_width=True,
        hide_index=True,
    )

    st.subheader("사용 모델")
    st.write("- Linear Regression\n- StandardScaler 기반 입력 표준화")

    st.subheader("현재 기능")
    st.write(
        """
        - 생활습관 기반 학업성과 예측
        - 개인별 민감도 분석
        - 미래 계획 시뮬레이션
        - 목표 기반 추천안 자동 탐색
        - 추천안 선택 및 카드 강조
        - 실행 난이도와 3주 단계계획
        - 일주일 실천계획과 진행률
        - 사용자용·분석용 시트를 구분한 엑셀 다운로드
        """
    )

    st.subheader("한계 및 주의사항")
    st.warning(
        """
        학업성과 환산 점수는 실제 대학의 GPA가 아닙니다.
        본 서비스는 학습 데이터의 통계적 패턴을 이용한 모의 분석 및 계획 지원 시스템이며,
        실제 성적이나 학업 결과를 보장하지 않습니다.
        """
    )
